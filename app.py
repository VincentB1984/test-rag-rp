'''
=============================================================
  RAG RECENSEMENT v3 — Serveur Web (FastAPI)
=============================================================
  Améliorations par rapport à la v1 :
  1. CHUNKING SÉMANTIQUE : découpage par page PDF, slide ODP,
     paragraphe ODT, feuille tableur — plus de coupures arbitraires.
  2. MÉTADONNÉES ENRICHIES : catégorie, sous-catégorie et année
     extraites automatiquement de la structure des dossiers.
  3. RERANKING : après récupération de 20 candidats, un Cross-Encoder
     (BAAI/bge-reranker-v2-m3) reclasse et sélectionne les 5 meilleurs.
  4. STREAMING : les tokens sont envoyés au fur et à mesure via
     Server-Sent Events (SSE), l'interface répond instantanément.

  VARIABLES D'ENVIRONNEMENT :
    ALBERT_API_KEY   : votre clé API Albert
    ALBERT_BASE_URL  : https://albert.api.etalab.gouv.fr/v1
    ALBERT_MODEL     : mistralai/Mistral-Small-3.2-24B-Instruct-2506
    EMBED_MODEL      : BAAI/bge-m3
    RERANK_MODEL     : BAAI/bge-reranker-v2-m3
    DOCS_DIR         : chemin vers le dossier de documents (défaut: ./documents)
    FAISS_INDEX      : nom du dossier de la base vectorielle
                       (défaut: faiss_index_recensement_v3)
=============================================================
'''

import os, sys, zipfile, re, threading, shutil, json, asyncio, math
from typing import List, Dict, Optional
from contextlib import asynccontextmanager

import requests as _requests
from langchain_core.embeddings import Embeddings as _LCEmbeddings

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from lxml import etree
from langchain_community.document_loaders import PyPDFLoader
from langchain_openai import ChatOpenAI
from langchain_community.vectorstores import FAISS
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import RunnablePassthrough, RunnableLambda
from langchain_core.output_parsers import StrOutputParser
from langchain_core.documents import Document


# ─────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────

def log(msg: str):
    print(msg, flush=True)
    sys.stdout.flush()


# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────

ALBERT_API_KEY  = os.getenv("ALBERT_API_KEY", "")
ALBERT_BASE_URL = os.getenv("ALBERT_BASE_URL", "https://albert.api.etalab.gouv.fr/v1")
ALBERT_MODEL    = os.getenv("ALBERT_MODEL", "mistralai/Mistral-Small-3.2-24B-Instruct-2506")
EMBED_MODEL     = os.getenv("EMBED_MODEL", "BAAI/bge-m3")
RERANK_MODEL    = os.getenv("RERANK_MODEL", "BAAI/bge-reranker-v2-m3")
DOCS_DIR        = os.getenv("DOCS_DIR", "./documents")
FAISS_INDEX     = os.getenv("FAISS_INDEX", "faiss_index_recensement_v3")
EXTENSIONS_SUPPORTEES = ('.pdf', '.odp', '.odt', '.xls', '.xlsx', '.ods')

for d in ["static", "documents", "templates"]:
    os.makedirs(d, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# EMBEDDINGS ALBERT
# ─────────────────────────────────────────────────────────────

class AlbertEmbeddings(_LCEmbeddings):
    def __init__(self, api_key: str, base_url: str, model: str):
        self._api_key = api_key
        self._base_url = base_url.rstrip("/")
        self._model = model
        self._headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    def _embed(self, texts: List[str]) -> List[List[float]]:
        all_embeddings = []
        for i in range(0, len(texts), 32):
            payload = {"model": self._model, "input": texts[i:i+32], "encoding_format": "float"}
            resp = _requests.post(f"{self._base_url}/embeddings", headers=self._headers, json=payload, timeout=120)
            if not resp.ok:
                raise RuntimeError(f"Albert API error {resp.status_code}: {resp.text[:500]}")
            items = sorted(resp.json()["data"], key=lambda x: x["index"])
            all_embeddings.extend(item["embedding"] for item in items)
        return all_embeddings

    def embed_documents(self, texts: List[str]) -> List[List[float]]:
        return self._embed(texts)

    def embed_query(self, text: str) -> List[float]:
        return self._embed([text])[0]


# ─────────────────────────────────────────────────────────────
# RERANKER (Cross-Encoder local)
# ─────────────────────────────────────────────────────────────

class Reranker:
    """
    Reclasse les documents récupérés par le retriever vectoriel
    en utilisant un modèle Cross-Encoder (BAAI/bge-reranker-v2-m3).
    Ce modèle évalue la pertinence de chaque paire (question, document)
    plutôt que de comparer des vecteurs séparément.
    """
    def __init__(self):
        self._model = None
        self._disponible = False

    def charger(self, model_name: str):
        try:
            from sentence_transformers import CrossEncoder
            self._model = CrossEncoder(model_name)
            self._disponible = True
            log(f"[INFO] Reranker chargé : {model_name}")
        except ImportError:
            log("[WARN] sentence-transformers non installé. Reranking désactivé.")
        except Exception as e:
            log(f"[WARN] Erreur chargement Reranker : {e}. Reranking désactivé.")

    def rerank(self, query: str, docs: List[Document], k: int = 5) -> List[Document]:
        if not self._disponible or not docs:
            return docs[:k]
        pairs = [[query, d.page_content] for d in docs]
        scores = self._model.predict(pairs)
        ranked = sorted(zip(scores, docs), key=lambda x: x[0], reverse=True)
        return [doc for _, doc in ranked[:k]]


# ─────────────────────────────────────────────────────────────
# AMÉLIORATION 1 & 2 : CHUNKING SÉMANTIQUE + MÉTADONNÉES ENRICHIES
# ─────────────────────────────────────────────────────────────

NS_DRAW = 'urn:oasis:names:tc:opendocument:xmlns:drawing:1.0'
NS_TEXT = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'


def enrichir_metadata(base_dir: str, file_path: str) -> Dict:
    """
    Extrait des métadonnées structurées depuis le chemin du fichier.
    Exemple : documents/02 - Formations/Module A/cours.pdf
    → category="02 - Formations", sub_category="Module A", year=2026 (si trouvé)
    """
    relative_path = os.path.relpath(file_path, base_dir).replace("\\", "/")
    parts = relative_path.split("/")
    meta = {"source": relative_path}
    if len(parts) > 1:
        meta["category"] = parts[0]
    if len(parts) > 2:
        meta["sub_category"] = "/".join(parts[1:-1])
    year_match = re.search(r'(20[2-9][0-9])', relative_path)
    if year_match:
        meta["year"] = int(year_match.group(1))
    return meta


def extraire_tout_texte(element) -> str:
    return ' '.join(p.strip() for p in element.itertext() if p.strip())


def charger_pdf_semantique(chemin: str, meta_base: Dict) -> List[Document]:
    """Chunking sémantique PDF : une page = un document."""
    docs = []
    try:
        pages = PyPDFLoader(chemin).load()
        for page in pages:
            if page.page_content.strip():
                page.metadata.update(meta_base)
                page.metadata["type"] = "pdf"
                docs.append(page)
    except Exception as e:
        log(f"  [WARN] PDF non lisible ({os.path.basename(chemin)}) : {e}")
    return docs


def charger_odp_semantique(chemin: str, meta_base: Dict) -> List[Document]:
    """Chunking sémantique ODP : une diapositive = un document."""
    docs = []
    try:
        with zipfile.ZipFile(chemin, 'r') as z:
            if 'content.xml' not in z.namelist():
                return []
            with z.open('content.xml') as f:
                tree = etree.parse(f)
        pages = tree.findall(f'.//{{{NS_DRAW}}}page')
        for i, page in enumerate(pages, start=1):
            texte = extraire_tout_texte(page).strip()
            if texte:
                meta = {**meta_base, "type": "odp", "slide": i}
                docs.append(Document(page_content=texte, metadata=meta))
    except Exception as e:
        log(f"  [WARN] ODP non lisible ({os.path.basename(chemin)}) : {e}")
    return docs


def charger_odt_semantique(chemin: str, meta_base: Dict) -> List[Document]:
    """
    Chunking sémantique ODT : découpage par paragraphes.
    Les paragraphes courts (<50 chars) sont fusionnés avec le suivant
    pour éviter des chunks trop petits.
    """
    docs = []
    try:
        with zipfile.ZipFile(chemin, 'r') as z:
            if 'content.xml' not in z.namelist():
                return []
            with z.open('content.xml') as f:
                tree = etree.parse(f)

        paragraphes_bruts = []
        for p in tree.findall(f'.//{{{NS_TEXT}}}p'):
            texte = extraire_tout_texte(p).strip()
            if texte:
                paragraphes_bruts.append(texte)

        # Fusion des paragraphes courts
        blocs, tampon = [], []
        for p in paragraphes_bruts:
            tampon.append(p)
            if len(' '.join(tampon)) >= 300:
                blocs.append(' '.join(tampon))
                tampon = []
        if tampon:
            blocs.append(' '.join(tampon))

        for i, bloc in enumerate(blocs, start=1):
            meta = {**meta_base, "type": "odt", "paragraph_block": i}
            docs.append(Document(page_content=bloc, metadata=meta))

    except Exception as e:
        log(f"  [WARN] ODT non lisible ({os.path.basename(chemin)}) : {e}")
    return docs


def charger_tableur_semantique(chemin: str, meta_base: Dict) -> List[Document]:
    """Chunking sémantique tableur : une feuille = un document."""
    docs = []
    nom = os.path.basename(chemin)
    ext = nom.lower().rsplit(".", 1)[-1]
    try:
        if ext == "xls":
            import xlrd
            wb = xlrd.open_workbook(chemin)
            for sheet in wb.sheets():
                lignes = []
                for row_idx in range(sheet.nrows):
                    cellules = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
                    ligne = "\t".join(c for c in cellules if c)
                    if ligne.strip():
                        lignes.append(ligne)
                contenu = "\n".join(lignes)
                if contenu.strip():
                    meta = {**meta_base, "type": "xls", "sheet": sheet.name}
                    docs.append(Document(page_content=contenu, metadata=meta))
        elif ext == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lignes = []
                for row in ws.iter_rows(values_only=True):
                    cellules = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cellules:
                        lignes.append("\t".join(cellules))
                contenu = "\n".join(lignes)
                if contenu.strip():
                    meta = {**meta_base, "type": "xlsx", "sheet": sheet_name}
                    docs.append(Document(page_content=contenu, metadata=meta))
            wb.close()
        elif ext == "ods":
            from odf.opendocument import load as odf_load
            from odf import teletype
            from odf.table import Table, TableRow, TableCell
            doc_odf = odf_load(chemin)
            sheets = doc_odf.spreadsheet.getElementsByType(Table)
            for sheet in sheets:
                sheet_name = sheet.getAttribute("name") or "Feuille"
                rows = sheet.getElementsByType(TableRow)
                lignes = []
                for row in rows:
                    cells = row.getElementsByType(TableCell)
                    cellules = [teletype.extractText(cell).strip() for cell in cells if teletype.extractText(cell).strip()]
                    if cellules:
                        lignes.append("\t".join(cellules))
                contenu = "\n".join(lignes)
                if contenu.strip():
                    meta = {**meta_base, "type": "ods", "sheet": sheet_name}
                    docs.append(Document(page_content=contenu, metadata=meta))
    except Exception as e:
        log(f"  [WARN] Tableur non lisible ({nom}) : {e}")
    return docs


def charger_dossier(dossier: str) -> List[Document]:
    """
    Charge récursivement tous les documents supportés.
    Applique le chunking sémantique et enrichit les métadonnées.
    """
    tous = []
    if not os.path.isdir(dossier):
        log(f"  [WARN] Dossier introuvable : {dossier}")
        return tous

    fichiers = []
    for racine, sous_dossiers, noms in os.walk(dossier):
        sous_dossiers.sort()
        for nom in sorted(noms):
            if nom.lower().endswith(EXTENSIONS_SUPPORTEES):
                fichiers.append(os.path.join(racine, nom))

    if not fichiers:
        log(f"  [WARN] Aucun fichier supporté dans : {dossier}")
        return tous

    log(f"  [INFO] {len(fichiers)} fichier(s) trouvé(s) dans {dossier} (récursif)")

    for chemin in fichiers:
        ext = os.path.splitext(chemin)[1].lower().lstrip(".")
        meta_base = enrichir_metadata(dossier, chemin)
        try:
            if ext == "pdf":
                docs = charger_pdf_semantique(chemin, meta_base)
            elif ext == "odp":
                docs = charger_odp_semantique(chemin, meta_base)
            elif ext == "odt":
                docs = charger_odt_semantique(chemin, meta_base)
            elif ext in ("xls", "xlsx", "ods"):
                docs = charger_tableur_semantique(chemin, meta_base)
            else:
                continue
            log(f"  [OK] {ext.upper()}: {meta_base['source']} → {len(docs)} chunk(s)")
            tous.extend(docs)
        except Exception as e:
            log(f"  [SKIP] {meta_base['source']} : {e}")

    return tous


# ─────────────────────────────────────────────────────────────
# ÉTAT GLOBAL DE L'APPLICATION
# ─────────────────────────────────────────────────────────────

class AppState:
    llm: Optional[ChatOpenAI] = None
    vectorstore: Optional[FAISS] = None
    reranker: Reranker = Reranker()
    pret: bool = False
    en_cours: bool = False
    message_init: str = ""
    nb_docs: int = 0
    nb_chunks: int = 0

state = AppState()


# ─────────────────────────────────────────────────────────────
# RECHERCHE WEB
# ─────────────────────────────────────────────────────────────

def recherche_web(question: str) -> str:
    try:
        from duckduckgo_search import DDGS
        with DDGS() as ddgs:
            resultats = list(ddgs.text(
                question + " site:legifrance.gouv.fr OR site:insee.fr OR site:service-public.fr",
                max_results=5
            ))
        if not resultats:
            with DDGS() as ddgs2:
                resultats = list(ddgs2.text(question, max_results=5))
        return "\n---\n".join(
            f"[{r.get('title','')}]\n{r.get('body','')}\nSource : {r.get('href','')}"
            for r in resultats
        )
    except Exception as e:
        return f"Recherche web indisponible : {e}"


# ─────────────────────────────────────────────────────────────
# AMÉLIORATION 3 : PIPELINE RAG AVEC RERANKING
# ─────────────────────────────────────────────────────────────

def run_rag_avec_reranking(question: str) -> str:
    """
    Pipeline complet :
    1. Récupère 20 candidats par similarité vectorielle
    2. Reclasse avec le Cross-Encoder → garde les 5 meilleurs
    3. Formate le contexte avec sources
    4. Génère la réponse via le LLM
    """
    candidats = state.vectorstore.similarity_search(question, k=20)
    meilleurs = state.reranker.rerank(question, candidats, k=5)

    def formater_source(d: Document) -> str:
        meta = d.metadata
        loc = meta.get("slide") or meta.get("sheet") or meta.get("page") or "?"
        type_loc = "slide" if meta.get("type") == "odp" else "feuille" if meta.get("type") in ("xls", "xlsx", "ods") else "page"
        return f"[{meta.get('source','?')}, {type_loc} {loc}]\n{d.page_content}"

    context = "\n---\n".join(formater_source(d) for d in meilleurs)

    prompt = ChatPromptTemplate.from_template(
        "Tu es un assistant expert du recensement de la population française.\n"
        "Ta mission est de fournir des réponses complètes, détaillées et structurées "
        "en te basant UNIQUEMENT sur les extraits de documents fournis.\n\n"
        "RÈGLES :\n"
        "1. Sois exhaustif dans ta réponse, structure-la avec des titres (###) et des listes.\n"
        "2. Mets en gras (**) les termes importants.\n"
        "3. Cite obligatoirement la source entre crochets après chaque information clé.\n"
        "4. Si les extraits ne contiennent pas la réponse, dis-le explicitement.\n"
        "5. Si la question demande une liste exhaustive (tous les noms, toutes les étapes...), "
        "précise que ta réponse est basée sur les extraits les plus pertinents, "
        "pas sur l'intégralité du corpus.\n\n"
        "Extraits de documents :\n{context}\n\n"
        "Question : {question}\n\n"
        "Réponse détaillée et structurée :"
    )
    chain = prompt | state.llm | StrOutputParser()
    return chain.invoke({"context": context, "question": question})


def run_rag_stream(question: str):
    """Même pipeline que run_rag_avec_reranking mais retourne un générateur de tokens."""
    candidats = state.vectorstore.similarity_search(question, k=20)
    meilleurs = state.reranker.rerank(question, candidats, k=5)

    def formater_source(d: Document) -> str:
        meta = d.metadata
        loc = meta.get("slide") or meta.get("sheet") or meta.get("page") or "?"
        type_loc = "slide" if meta.get("type") == "odp" else "feuille" if meta.get("type") in ("xls", "xlsx", "ods") else "page"
        return f"[{meta.get('source','?')}, {type_loc} {loc}]\n{d.page_content}"

    context = "\n---\n".join(formater_source(d) for d in meilleurs)

    prompt = ChatPromptTemplate.from_template(
        "Tu es un assistant expert du recensement de la population française.\n"
        "Ta mission est de fournir des réponses complètes, détaillées et structurées "
        "en te basant UNIQUEMENT sur les extraits de documents fournis.\n\n"
        "RÈGLES :\n"
        "1. Sois exhaustif dans ta réponse, structure-la avec des titres (###) et des listes.\n"
        "2. Mets en gras (**) les termes importants.\n"
        "3. Cite obligatoirement la source entre crochets après chaque information clé.\n"
        "4. Si les extraits ne contiennent pas la réponse, dis-le explicitement.\n"
        "5. Si la question demande une liste exhaustive (tous les noms, toutes les étapes...), "
        "précise que ta réponse est basée sur les extraits les plus pertinents, "
        "pas sur l'intégralité du corpus.\n\n"
        "Extraits de documents :\n{context}\n\n"
        "Question : {question}\n\n"
        "Réponse détaillée et structurée :"
    )
    chain = prompt | state.llm | StrOutputParser()
    return chain.stream({"context": context, "question": question})


def run_web(question: str) -> str:
    context = recherche_web(question)
    prompt = ChatPromptTemplate.from_template(
        "Tu es un assistant expert du recensement de la population française.\n"
        "Réponds en te basant sur les résultats de recherche web fournis. Cite les sources (URLs).\n\n"
        "Résultats web :\n{context}\n\nQuestion : {question}\nRéponse :"
    )
    chain = prompt | state.llm | StrOutputParser()
    return chain.invoke({"context": context, "question": question})


def run_web_stream(question: str):
    context = recherche_web(question)
    prompt = ChatPromptTemplate.from_template(
        "Tu es un assistant expert du recensement de la population française.\n"
        "Réponds en te basant sur les résultats de recherche web fournis. Cite les sources (URLs).\n\n"
        "Résultats web :\n{context}\n\nQuestion : {question}\nRéponse :"
    )
    chain = prompt | state.llm | StrOutputParser()
    return chain.stream({"context": context, "question": question})


def run_llm_stream(question: str):
    prompt = ChatPromptTemplate.from_template(
        "Tu es un assistant expert du recensement de la population française et de la rédaction administrative.\n"
        "Réponds en français de manière professionnelle et structurée.\n\n"
        "Question : {question}\nRéponse :"
    )
    chain = prompt | state.llm | StrOutputParser()
    return chain.stream({"question": question})


def router(question: str) -> str:
    prompt = ChatPromptTemplate.from_template(
        "Analyse cette question et réponds UNIQUEMENT par un seul mot parmi : RAG, WEB, LLM\n\n"
        "- RAG : question sur des procédures internes, formations, consignes, tournées, questionnaires, rôles des agents\n"
        "- WEB : question sur des décrets, lois, dates officielles, actualités, textes réglementaires\n"
        "- LLM : rédaction de courriers, reformulations, calculs, questions générales\n\n"
        "Question : {question}\nRéponse (un seul mot) :"
    )
    chain = prompt | state.llm | StrOutputParser()
    try:
        mode_brut = chain.invoke({"question": question}).strip().upper()
        return "RAG" if "RAG" in mode_brut else "WEB" if "WEB" in mode_brut else "LLM"
    except Exception:
        return "RAG"


# ─────────────────────────────────────────────────────────────
# INITIALISATION DU RAG
# ─────────────────────────────────────────────────────────────

def _construire_rag(forcer: bool = False):
    log("\n" + "="*60)
    log("  RAG RECENSEMENT v3 — Initialisation...")
    log("="*60)
    state.pret, state.en_cours, state.message_init = False, True, "Construction en cours..."

    try:
        if not ALBERT_API_KEY:
            raise ValueError("ALBERT_API_KEY non définie dans les variables d'environnement.")

        log(f"[INFO] LLM : {ALBERT_MODEL}")
        state.llm = ChatOpenAI(
            model=ALBERT_MODEL, temperature=0,
            api_key=ALBERT_API_KEY, base_url=ALBERT_BASE_URL
        )

        log(f"[INFO] Embeddings : {EMBED_MODEL}")
        embeddings = AlbertEmbeddings(
            api_key=ALBERT_API_KEY, base_url=ALBERT_BASE_URL, model=EMBED_MODEL
        )

        log(f"[INFO] Reranker : {RERANK_MODEL}")
        state.reranker.charger(RERANK_MODEL)

        # Reconstitution depuis les morceaux (si index splitté pour GitHub)
        def reconstituer_depuis_parts(fichier_cible: str):
            manifest = fichier_cible + ".manifest"
            if os.path.exists(manifest) and not os.path.exists(fichier_cible):
                with open(manifest) as mf:
                    nb_parts = int(mf.read().strip())
                log(f"[INFO] Reconstitution de {os.path.basename(fichier_cible)} ({nb_parts} morceaux)...")
                with open(fichier_cible, "wb") as out:
                    for i in range(nb_parts):
                        part_path = f"{fichier_cible}.part{i:02d}"
                        if not os.path.exists(part_path):
                            raise FileNotFoundError(f"Morceau manquant : {part_path}")
                        with open(part_path, "rb") as pf:
                            while True:
                                bloc = pf.read(4 * 1024 * 1024)
                                if not bloc: break
                                out.write(bloc)
                        os.remove(part_path)
                os.remove(manifest)
                log(f"[INFO] Reconstitution terminée.")

        faiss_file = os.path.join(FAISS_INDEX, "index.faiss")
        pkl_file   = os.path.join(FAISS_INDEX, "index.pkl")
        reconstituer_depuis_parts(faiss_file)
        reconstituer_depuis_parts(pkl_file)

        if not forcer and os.path.exists(faiss_file) and os.path.exists(pkl_file):
            log(f"[INFO] Chargement de la base vectorielle : {FAISS_INDEX}")
            state.vectorstore = FAISS.load_local(
                FAISS_INDEX, embeddings, allow_dangerous_deserialization=True
            )
            state.nb_chunks = state.vectorstore.index.ntotal
            state.nb_docs = len(set(
                d.metadata.get("source", "") for d in state.vectorstore.docstore._dict.values()
            ))
            log(f"[INFO] Base chargée : {state.nb_chunks} chunks, {state.nb_docs} documents sources.")
        else:
            log(f"[INFO] Construction depuis les documents : {DOCS_DIR}")
            documents = charger_dossier(DOCS_DIR)
            if not documents:
                state.message_init = "Aucun document trouvé. Uploadez des fichiers via /admin."
                log(f"[WARN] {state.message_init}")
                state.en_cours = False
                return

            state.nb_docs = len(set(d.metadata.get("source", "") for d in documents))
            state.nb_chunks = len(documents)
            log(f"[INFO] {state.nb_chunks} chunks créés, vectorisation en cours...")

            if os.path.exists(FAISS_INDEX):
                shutil.rmtree(FAISS_INDEX)
            state.vectorstore = FAISS.from_documents(documents, embeddings)
            state.vectorstore.save_local(FAISS_INDEX)
            log(f"[INFO] Base FAISS sauvegardée : {FAISS_INDEX}")

        state.pret = True
        state.message_init = f"Prêt — {state.nb_docs} documents, {state.nb_chunks} chunks indexés."
        log(f"\n[OK] {state.message_init}\n")

    except Exception as e:
        import traceback
        state.message_init = f"Erreur d'initialisation : {e}"
        log(f"[ERREUR] {state.message_init}")
        log(traceback.format_exc())
        state.pret = False
    finally:
        state.en_cours = False


def initialiser_rag_background(forcer: bool = False):
    if state.en_cours:
        log("[INFO] Initialisation déjà en cours.")
        return
    threading.Thread(target=_construire_rag, args=(forcer,), daemon=True).start()


# ─────────────────────────────────────────────────────────────
# APPLICATION FASTAPI
# ─────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(application: FastAPI):
    log("[INFO] Démarrage FastAPI — lancement de l'initialisation RAG...")
    initialiser_rag_background(forcer=False)
    yield
    log("[INFO] Arrêt de l'application.")


app = FastAPI(title="RAG Recensement v3", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory="static"), name="static")


class MessageRequest(BaseModel):
    question: str
    mode_force: str = "AUTO"

class MessageResponse(BaseModel):
    reponse: str
    mode: str
    sources: List[str] = []


@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse("templates/index.html")


@app.get("/health")
async def health():
    return {
        "status": "ok" if state.pret else ("construction" if state.en_cours else "attente"),
        "message": state.message_init,
        "nb_chunks": state.nb_chunks,
        "nb_docs": state.nb_docs,
    }


@app.get("/healthz")
async def healthz():
    return {"status": "ok"}


# ─────────────────────────────────────────────────────────────
# AMÉLIORATION 4 : STREAMING (Server-Sent Events)
# ─────────────────────────────────────────────────────────────

@app.post("/chat/stream")
async def chat_stream(req: MessageRequest):
    """
    Endpoint de streaming : envoie les tokens au fur et à mesure.
    Format : Server-Sent Events (SSE).
    Chaque événement : data: {"token": "..."}\n\n
    Fin de stream : data: {"end": true, "mode": "RAG"}\n\n
    """
    if not state.pret:
        detail = "Le RAG est en cours de construction." if state.en_cours else "Le RAG n'est pas initialisé."
        raise HTTPException(status_code=503, detail=detail)

    question = req.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="La question ne peut pas être vide.")

    mode = req.mode_force.upper()
    if mode == "AUTO":
        mode = await asyncio.to_thread(router, question)

    async def event_generator():
        try:
            if mode == "RAG":
                gen = await asyncio.to_thread(run_rag_stream, question)
            elif mode == "WEB":
                gen = await asyncio.to_thread(run_web_stream, question)
            else:
                gen = await asyncio.to_thread(run_llm_stream, question)

            for chunk in gen:
                yield f"data: {json.dumps({'token': chunk}, ensure_ascii=False)}\n\n"
                await asyncio.sleep(0)  # Cède le contrôle à l'event loop

        except Exception as e:
            log(f"[ERREUR STREAM] {e}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        finally:
            yield f"data: {json.dumps({'end': True, 'mode': mode})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.post("/chat", response_model=MessageResponse)
async def chat(req: MessageRequest):
    """Endpoint classique (non-streaming) — conservé pour compatibilité."""
    if not state.pret:
        detail = "Le RAG est en cours de construction." if state.en_cours else "Le RAG n'est pas initialisé."
        raise HTTPException(status_code=503, detail=detail)

    question = req.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="La question ne peut pas être vide.")

    mode = req.mode_force.upper()
    if mode == "AUTO":
        mode = await asyncio.to_thread(router, question)

    try:
        if mode == "RAG":
            reponse = await asyncio.to_thread(run_rag_avec_reranking, question)
        elif mode == "WEB":
            reponse = await asyncio.to_thread(run_web, question)
        else:
            prompt = ChatPromptTemplate.from_template(
                "Tu es un expert du recensement. Réponds professionnellement.\nQuestion: {question}\nRéponse:"
            )
            chain = prompt | state.llm | StrOutputParser()
            reponse = await asyncio.to_thread(chain.invoke, {"question": question})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération : {e}")

    return MessageResponse(reponse=reponse, mode=mode)


# ─────────────────────────────────────────────────────────────
# ROUTES D'ADMINISTRATION
# ─────────────────────────────────────────────────────────────

@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    fichiers = []
    if os.path.isdir(DOCS_DIR):
        for racine, sous_dossiers, noms in os.walk(DOCS_DIR):
            sous_dossiers.sort()
            for nom in sorted(noms):
                if nom.lower().endswith(EXTENSIONS_SUPPORTEES):
                    chemin_relatif = os.path.relpath(os.path.join(racine, nom), DOCS_DIR)
                    fichiers.append(chemin_relatif.replace("\\", "/"))

    statut_rag = (
        f'<span style="color:#10b981">OK Pret -- {state.nb_docs} docs, {state.nb_chunks} chunks</span>'
        if state.pret else
        '<span style="color:#f59e0b">Construction en cours...</span>'
        if state.en_cours else
        f'<span style="color:#ef4444">Non initialise -- {state.message_init}</span>'
    )

    liste_fichiers = "".join(
        f'<li style="display:flex;justify-content:space-between;align-items:center;'
        f'padding:6px 0;border-bottom:1px solid #eee">'
        f'<span>{f}</span>'
        f'<button onclick="supprimerFichier(\'{f}\')" '
        f'style="background:#ef4444;color:#fff;border:none;border-radius:4px;'
        f'padding:2px 8px;cursor:pointer;font-size:12px">Supprimer</button>'
        f'</li>'
        for f in fichiers
    ) or "<li style='color:#6b7280'>Aucun document charge</li>"

    html = f'''<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Administration -- RAG Recensement v3</title>
  <style>
    body {{ font-family: "Segoe UI", Arial, sans-serif; background:#f5f6fa; color:#1a1a2e; margin:0; padding:0; }}
    header {{ background:#003189; color:#fff; padding:16px 24px; display:flex; align-items:center; gap:12px; }}
    header h1 {{ font-size:18px; margin:0; }}
    .container {{ max-width:700px; margin:32px auto; padding:0 16px; }}
    .card {{ background:#fff; border-radius:12px; padding:24px; box-shadow:0 2px 12px rgba(0,0,0,0.08); margin-bottom:20px; }}
    h2 {{ font-size:15px; font-weight:700; margin:0 0 16px; text-transform:uppercase; letter-spacing:0.5px; color:#6b7280; }}
    .statut {{ padding:12px 16px; border-radius:8px; background:#f0f4ff; border:1px solid #c7d2fe; font-size:14px; margin-bottom:16px; }}
    .drop-zone {{ border:2px dashed #c7d2fe; border-radius:8px; padding:32px; text-align:center; color:#6b7280; cursor:pointer; background:#fafbff; }}
    .drop-zone:hover {{ border-color:#003189; background:#e8edf8; color:#003189; }}
    .drop-zone input {{ display:none; }}
    .btn {{ display:inline-block; padding:10px 20px; border-radius:8px; border:none; cursor:pointer; font-size:14px; font-weight:600; }}
    .btn-success {{ background:#10b981; color:#fff; }}
    ul {{ list-style:none; padding:0; margin:0; }}
    #log {{ background:#1a1a2e; color:#a3e635; font-family:monospace; font-size:12px; padding:12px; border-radius:8px; min-height:60px; max-height:200px; overflow-y:auto; white-space:pre-wrap; display:none; margin-top:12px; }}
    #progression {{ margin-top:12px; font-size:13px; color:#6b7280; }}
    .badge {{ display:inline-block; background:#003189; color:#fff; border-radius:4px; padding:2px 8px; font-size:11px; margin-left:8px; }}
  </style>
</head>
<body>
<header>
  <div style="width:4px;height:32px;background:linear-gradient(to bottom,#002395 33%,#fff 33%,#fff 66%,#e1000f 66%);border-radius:2px"></div>
  <h1>Administration -- RAG Recensement <span class="badge">v3</span></h1>
  <a href="/" style="margin-left:auto;color:#fff;font-size:13px">Retour au chat</a>
</header>
<div class="container">
  <div class="card">
    <h2>Statut du RAG</h2>
    <div class="statut">{statut_rag}</div>
    <div style="font-size:12px;color:#6b7280;margin-bottom:12px">
      Améliorations actives : Chunking sémantique | Métadonnées enrichies | Reranking | Streaming
    </div>
    <button class="btn btn-success" onclick="reconstruire()">Reconstruire le RAG</button>
    <div id="log"></div>
  </div>
  <div class="card">
    <h2>Uploader des documents</h2>
    <div class="drop-zone" id="dropZone" onclick="document.getElementById('fileInput').click()"
         ondragover="event.preventDefault()" ondrop="gererDrop(event)">
      <div style="font-size:36px;margin-bottom:8px">&#128193;</div>
      <div style="font-size:15px;font-weight:600;margin-bottom:4px">Cliquez ou glissez vos fichiers ici</div>
      <div style="font-size:12px">Formats : PDF, ODP, ODT, XLS, XLSX, ODS</div>
      <input type="file" id="fileInput" multiple accept=".pdf,.odp,.odt,.xls,.xlsx,.ods" onchange="uploaderFichiers(this.files)">
    </div>
    <div id="progression"></div>
  </div>
  <div class="card">
    <h2>Documents charges ({len(fichiers)} fichier(s))</h2>
    <ul id="listeFichiers">{liste_fichiers}</ul>
  </div>
</div>
<script>
  async function uploaderFichiers(files) {{
    if (!files || files.length === 0) return;
    const prog = document.getElementById('progression');
    prog.textContent = 'Envoi de ' + files.length + ' fichier(s)...';
    const form = new FormData();
    for (const f of files) form.append('files', f);
    try {{
      const r = await fetch('/admin/upload', {{ method: 'POST', body: form }});
      const data = await r.json();
      if (r.ok) {{
        prog.innerHTML = '<span style="color:#10b981">OK ' + data.message + '</span>';
        setTimeout(() => location.reload(), 1500);
      }} else {{
        prog.innerHTML = '<span style="color:#ef4444">Erreur ' + (data.detail || 'inconnue') + '</span>';
      }}
    }} catch(e) {{
      prog.innerHTML = '<span style="color:#ef4444">Erreur reseau</span>';
    }}
  }}
  function gererDrop(e) {{
    e.preventDefault();
    uploaderFichiers(e.dataTransfer.files);
  }}
  async function supprimerFichier(nom) {{
    if (!confirm('Supprimer "' + nom + '" ?')) return;
    const r = await fetch('/admin/supprimer', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{ nom }})
    }});
    if (r.ok) location.reload();
    else alert('Erreur');
  }}
  async function reconstruire() {{
    const log = document.getElementById('log');
    log.style.display = 'block';
    log.textContent = 'Lancement...\n';
    const r = await fetch('/admin/reconstruire', {{ method: 'POST' }});
    const data = await r.json();
    log.textContent += data.message + '\n';
    const poll = setInterval(async () => {{
      try {{
        const s = await fetch('/health');
        const sd = await s.json();
        log.textContent += sd.status + ' -- ' + (sd.message || sd.nb_chunks + ' chunks') + '\n';
        log.scrollTop = log.scrollHeight;
        if (sd.status === 'ok') {{
          clearInterval(poll);
          log.textContent += 'Termine !\n';
          setTimeout(() => location.reload(), 1500);
        }} else if (sd.status === 'attente') {{
          clearInterval(poll);
          setTimeout(() => location.reload(), 2000);
        }}
      }} catch(e) {{
        log.textContent += 'Erreur reseau...\n';
      }}
    }}, 5000);
  }}
</script>
</body>
</html>'''
    return HTMLResponse(html)


@app.post("/admin/upload")
async def upload_documents(files: List[UploadFile] = File(...)):
    os.makedirs(DOCS_DIR, exist_ok=True)
    extensions_valides = {ext.lstrip('.') for ext in EXTENSIONS_SUPPORTEES}
    sauvegardes = []
    for f in files:
        nom = f.filename or "fichier_inconnu"
        ext = nom.lower().rsplit(".", 1)[-1] if "." in nom else ""
        if ext not in extensions_valides:
            continue
        dest = os.path.join(DOCS_DIR, nom)
        contenu = await f.read()
        with open(dest, "wb") as out:
            out.write(contenu)
        sauvegardes.append(nom)
        log(f"[UPLOAD] {dest} ({len(contenu)} octets)")

    if not sauvegardes:
        raise HTTPException(
            status_code=400,
            detail=f"Aucun fichier valide. Formats : {', '.join(sorted(extensions_valides)).upper()}"
        )
    initialiser_rag_background(forcer=True)
    return JSONResponse({"message": f"{len(sauvegardes)} fichier(s) uploadé(s). Reconstruction lancée."})


@app.post("/admin/supprimer")
async def supprimer_document(payload: dict):
    nom = payload.get("nom", "")
    if not nom:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide.")
    chemin = os.path.normpath(os.path.join(DOCS_DIR, nom))
    if not chemin.startswith(os.path.normpath(DOCS_DIR)):
        raise HTTPException(status_code=400, detail="Chemin non autorisé.")
    if not os.path.exists(chemin):
        raise HTTPException(status_code=404, detail="Fichier introuvable.")
    os.remove(chemin)
    log(f"[ADMIN] Supprimé : {chemin}")
    return JSONResponse({"message": f"{nom} supprimé."})


@app.post("/admin/reconstruire")
async def reconstruire_rag():
    if state.en_cours:
        return JSONResponse({"message": "Construction déjà en cours, patientez."})
    initialiser_rag_background(forcer=True)
    return JSONResponse({"message": "Reconstruction du RAG lancée en arrière-plan."})


# ─────────────────────────────────────────────────────────────
# POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("app:app", host="0.0.0.0", port=port)
